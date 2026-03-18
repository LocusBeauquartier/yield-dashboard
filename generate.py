"""
generate.py — Générateur de dashboard Yield Beauquartier
=========================================================
Usage local :
    python generate.py --local --input ./data/ --output ./index.html

Usage GitHub Actions (SharePoint) :
    python generate.py --sharepoint (les fichiers sont déjà téléchargés dans ./data/)

Structure attendue dans ./data/ :
    export_J.xlsx           → Export J (onglet "Données")
    export_J1.xlsx          → Export J-1
    fenetre_J7.xlsx         → Fenêtre J-7
    fenetre_J14.xlsx        → Fenêtre J-14
    fenetre_J21.xlsx        → Fenêtre J-21 (proxy J-30 dans le dashboard)
    fenetre_J45.xlsx        → Fenêtre J-45
    budget.xlsx             → Budget (col A=mois, B=CA, C=PM, D=nuits, E=TO, F=RevPAR)
    reservations.xlsx       → Rapport réservations (onglet "Réservations")
    pickup_*.xlsx           → Exports montée en charge (plusieurs fichiers, date dans le nom)
"""

import argparse
import base64
import json
import os
import re
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl manquant — pip install openpyxl")
    sys.exit(1)

# ── CONFIG NOMS DE FICHIERS ────────────────────────────────────────────────────
FILE_MAP = {
    "main":       "export_J.xlsx",
    "j1":         "export_J1.xlsx",
    "fenetre_J-1":  "fenetre_J1.xlsx",
    "fenetre_J-3":  "fenetre_J3.xlsx",
    "fenetre_J-7":  "fenetre_J7.xlsx",
    "fenetre_J-14": "fenetre_J14.xlsx",
    "fenetre_J-21": "fenetre_J21.xlsx",
    "fenetre_J-45": "fenetre_J45.xlsx",
    "budget":     "budget.xlsx",
    "segtar":     "reservations.xlsx",
}
PICKUP_PATTERN = "pickup_*.xlsx"  # ou "export_*.xlsx" selon votre convention

MONTHS_FR = ["Jan","Fév","Mar","Avr","Mai","Jun","Jul","Aoû","Sep","Oct","Nov","Déc"]
MONTHS_FULL = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août",
               "Septembre","Octobre","Novembre","Décembre"]

# ── PARSEURS ───────────────────────────────────────────────────────────────────

def parse_date_cell(val):
    """Convertit une cellule date Excel en objet date Python."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val if isinstance(val, datetime) else datetime.combine(val, datetime.min.time())
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                pass
    return None


def parse_main_xlsx(path):
    """
    Parse un export principal (onglet 'Données').
    L1=dates, L2=TO, L5=occupés, L10=RevPAR, L11=PM, L14=CA
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Données" not in wb.sheetnames:
        raise ValueError(f"Onglet 'Données' introuvable dans {path}")
    ws = wb["Données"]
    rows = list(ws.iter_rows(values_only=True))

    date_row  = rows[0]   # L1 — dates
    to_row    = rows[1]   # L2 — TO
    occ_row   = rows[4]   # L5 — occupés
    revpar_row= rows[9]   # L10 — RevPAR
    pm_row    = rows[10]  # L11 — PM
    ca_row    = rows[13]  # L14 — CA

    result = []
    for c in range(3, len(date_row)):
        rd = date_row[c]
        if rd is None or rd == "Total":
            break
        dt = parse_date_cell(rd)
        if dt is None:
            continue
        result.append({
            "date": dt.strftime("%Y-%m-%dT00:00:00.000Z"),
            "ca":       float(ca_row[c]   or 0),
            "to":       float(to_row[c]   or 0),
            "pm":       float(pm_row[c]   or 0),
            "occupied": float(occ_row[c]  or 0),
            "revpar":   float(revpar_row[c] or 0),
        })
    wb.close()
    return result


def parse_budget_xlsx(path):
    """
    Parse le fichier budget.
    Col A=mois, B=CA, C=PM, D=nuits, E=TO, F=RevPAR
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    result = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        col_a = row[0]
        idx = -1
        if isinstance(col_a, (datetime, date)):
            idx = (col_a.month if isinstance(col_a, date) else col_a.month) - 1
        elif isinstance(col_a, str):
            ml = col_a.strip().lower()
            for i, m in enumerate(MONTHS_FULL):
                if ml.startswith(m.lower()):
                    idx = i
                    break
            if idx < 0:
                for i, m in enumerate(MONTHS_FR):
                    if ml.startswith(m.lower()):
                        idx = i
                        break
            if idx < 0:
                try:
                    n = int(ml)
                    if 1 <= n <= 12:
                        idx = n - 1
                except ValueError:
                    pass
        if idx < 0:
            continue
        def floatv(v):
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).replace(" ", "").replace(",", "."))
            except:
                return 0.0
        ca     = floatv(row[1]) if len(row) > 1 else 0
        pm     = floatv(row[2]) if len(row) > 2 else 0
        nights = floatv(row[3]) if len(row) > 3 else 0
        to     = floatv(row[4]) if len(row) > 4 else 0
        revpar = floatv(row[5]) if len(row) > 5 else 0
        if ca > 0 or pm > 0:
            result.append({"month": idx, "budget": ca, "pm": pm,
                           "nights": nights, "to": to, "revpar": revpar})
    wb.close()
    return result


def parse_segtar_xlsx(path):
    """
    Parse le rapport de réservations (onglet 'Réservations').
    Extrait : canaux (col AD=29), tarifs (col AE=30), nationalités (col H=7)
    Col Q=16 : arrivée, Col S=18 : nuits, Col AK=36 : montant, Col K=10 : statut
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws_name = "Réservations" if "Réservations" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))

    segtar_raw = []
    nat_raw    = []

    for row in rows[1:]:
        if not row or len(row) < 2:
            continue
        # Numéro non numérique = ligne Total
        try:
            float(row[0])
        except (TypeError, ValueError):
            continue
        statut = str(row[10] or "").strip()
        if statut == "Annulé":
            continue

        # Mois d'arrivée
        arrival = row[16]
        month = None
        if arrival is not None:
            if isinstance(arrival, (datetime, date)):
                month = (arrival.month if isinstance(arrival, date) else arrival.month) - 1
            elif isinstance(arrival, str):
                m = re.search(r"(\d{4})-(\d{2})-(\d{2})", arrival)
                if m:
                    month = int(m.group(2)) - 1
        if month is None:
            continue

        nights = float(row[18] or 0) if len(row) > 18 else 0
        ca     = float(row[36] or 0) if len(row) > 36 else 0
        seg    = str(row[29] or "").strip() or "Non renseigné" if len(row) > 29 else "Non renseigné"
        tar    = str(row[30] or "").strip() if len(row) > 30 else ""
        nat    = str(row[7]  or "").strip() or "N/C" if len(row) > 7 else "N/C"

        segtar_raw.append({"month": month, "seg": seg, "tar": tar,
                           "nights": nights, "ca": ca})
        nat_raw.append({"month": month, "nat": nat, "nights": nights, "ca": ca})

    wb.close()
    return segtar_raw, nat_raw


def extract_date_from_filename(name):
    """Extrait une date depuis le nom de fichier (YY.MM.DD, YYYY-MM-DD, DD.MM.YYYY…)"""
    stem = Path(name).stem
    patterns = [
        (r"(\d{4})-(\d{2})-(\d{2})", lambda m: datetime(int(m[1]), int(m[2]), int(m[3]))),
        (r"(\d{2})\.(\d{2})\.(\d{4})", lambda m: datetime(int(m[3]), int(m[2]), int(m[1]))),
        (r"(\d{2})\.(\d{2})\.(\d{2})(?!\d)", lambda m: datetime(
            2000 + int(m[1]) if int(m[1]) < 50 else 1900 + int(m[1]),
            int(m[2]), int(m[3]))),
        (r"(\d{4})(\d{2})(\d{2})", lambda m: datetime(int(m[1]), int(m[2]), int(m[3]))),
    ]
    for pat, fn in patterns:
        match = re.search(pat, stem)
        if match:
            try:
                return fn(match.groups())
            except ValueError:
                pass
    return None


def parse_pickup_xlsx(path):
    """Parse un export pickup (même format que export_J)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws_name = "Données" if "Données" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))
    date_row = rows[0]
    occ_row  = rows[4]
    nights_by_month = {}
    for c in range(3, len(date_row)):
        rd = date_row[c]
        if rd is None or rd == "Total":
            break
        dt = parse_date_cell(rd)
        if dt is None:
            continue
        m = dt.month - 1
        occ = float(occ_row[c] or 0) if occ_row and c < len(occ_row) else 0
        nights_by_month[m] = nights_by_month.get(m, 0) + occ
    wb.close()
    return nights_by_month


# ── INJECTION DANS LE HTML ─────────────────────────────────────────────────────

def inject_into_html(template_path, output_path, payload, export_date_str):
    """
    Lit le template HTML, remplace le bloc DEMO_DATA_START…END
    par les données réelles, écrit le résultat.
    """
    with open(template_path, "r", encoding="utf-8") as f:
        src = f.read()

    START = "/*DEMO_DATA_START*/"
    END   = "/*DEMO_DATA_END*/"
    i0 = src.find(START)
    i1 = src.find(END)
    if i0 < 0 or i1 < 0:
        raise ValueError("Marqueurs DEMO_DATA_START / DEMO_DATA_END introuvables dans le template HTML")

    snapshot_block = f"""{START}
(function(){{
  const P={json.dumps(payload, ensure_ascii=False, separators=(',', ':'))};
  const rv=s=>new Date(s);
  rawJ        = P.rawJ.map(r=>{{...r,date:rv(r.date)}});
  rawJ1       = P.rawJ1.map(r=>{{...r,date:rv(r.date)}});
  fenetreData = Object.fromEntries(Object.entries(P.fenetreData).map(([k,v])=>[k,v.map(r=>{{...r,date:rv(r.date)}})]));
  budgetData  = P.budgetData;
  canauxData  = P.canauxData;
  tarifsData  = P.tarifsData;
  segtarRaw   = P.segtarRaw;
  natRaw      = P.natRaw;
  pickupSnapshots = P.pickupSnapshots.map(s=>{{...s,exportDate:rv(s.exportDate)}});
  monthly     = agg(rawJ);
  const ig=document.getElementById('importsGrid'); if(ig)ig.style.display='none';
  const eb=document.querySelector('[onclick="exportSnapshot()"]'); if(eb)eb.style.display='none';
  document.getElementById('sub').textContent='Rapport généré le {export_date_str}';
}})();
{END}"""

    html = src[:i0] + snapshot_block + src[i1 + len(END):]

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"✅ Dashboard généré → {output_path}")


# ── MAIN ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Générateur de dashboard Yield Beauquartier")
    parser.add_argument("--input",    default="./data",         help="Dossier contenant les xlsx")
    parser.add_argument("--template", default="./RMS_MSC.html", help="Fichier template HTML")
    parser.add_argument("--output",   default="./index.html",   help="Fichier de sortie")
    args = parser.parse_args()

    data_dir   = Path(args.input)
    template   = Path(args.template)
    output     = Path(args.output)

    if not template.exists():
        print(f"❌ Template introuvable : {template}")
        sys.exit(1)

    payload = {
        "rawJ":       [],
        "rawJ1":      [],
        "fenetreData": {},
        "budgetData": [],
        "canauxData": [],
        "tarifsData": [],
        "segtarRaw":  [],
        "natRaw":     [],
        "pickupSnapshots": [],
    }

    errors = []

    # ── Export J (principal) ──
    f = data_dir / FILE_MAP["main"]
    if f.exists():
        try:
            payload["rawJ"] = parse_main_xlsx(f)
            print(f"✓ Export J       — {len(payload['rawJ'])} jours")
        except Exception as e:
            errors.append(f"Export J : {e}")
    else:
        print(f"⚠ Export J       — fichier absent ({f})")

    # ── Export J-1 ──
    f = data_dir / FILE_MAP["j1"]
    if f.exists():
        try:
            payload["rawJ1"] = parse_main_xlsx(f)
            print(f"✓ Export J-1     — {len(payload['rawJ1'])} jours")
        except Exception as e:
            errors.append(f"Export J-1 : {e}")
    else:
        print(f"⚠ Export J-1     — fichier absent ({f})")

    # ── Fenêtres ──
    for key in ["J-1", "J-3", "J-7", "J-14", "J-21", "J-45"]:
        fkey = f"fenetre_{key}"
        if fkey not in FILE_MAP:
            continue
        f = data_dir / FILE_MAP[fkey]
        if f.exists():
            try:
                rows = parse_main_xlsx(f)
                payload["fenetreData"][key] = rows
                print(f"✓ Fenêtre {key:<5}  — {len(rows)} jours")
            except Exception as e:
                errors.append(f"Fenêtre {key} : {e}")
        else:
            print(f"⚠ Fenêtre {key:<5}  — fichier absent ({f})")

    # ── Budget ──
    f = data_dir / FILE_MAP["budget"]
    if f.exists():
        try:
            payload["budgetData"] = parse_budget_xlsx(f)
            print(f"✓ Budget         — {len(payload['budgetData'])} mois")
        except Exception as e:
            errors.append(f"Budget : {e}")
    else:
        print(f"⚠ Budget         — fichier absent ({f})")

    # ── Réservations (segments + tarifs + nationalités) ──
    f = data_dir / FILE_MAP["segtar"]
    if f.exists():
        try:
            segtar_raw, nat_raw = parse_segtar_xlsx(f)
            payload["segtarRaw"] = segtar_raw
            payload["natRaw"]    = nat_raw
            print(f"✓ Réservations   — {len(segtar_raw)} lignes, {len(nat_raw)} nationalités")
        except Exception as e:
            errors.append(f"Réservations : {e}")
    else:
        print(f"⚠ Réservations   — fichier absent ({f})")

    # ── Pickups (montée en charge) ──
    pickup_files = sorted(data_dir.glob(PICKUP_PATTERN))
    seen_dates = {}
    for pf in pickup_files:
        export_date = extract_date_from_filename(pf.name)
        if export_date is None:
            print(f"⚠ Pickup ignoré  — date non trouvée dans {pf.name}")
            continue
        key = export_date.strftime("%Y-%m-%d")
        if key in seen_dates:
            continue
        seen_dates[key] = True
        try:
            nights_by_month = parse_pickup_xlsx(pf)
            payload["pickupSnapshots"].append({
                "exportDate": export_date.strftime("%Y-%m-%dT00:00:00.000Z"),
                "name":       pf.name,
                "nightsByMonth": {str(k): v for k, v in nights_by_month.items()},
            })
            print(f"✓ Pickup         — {pf.name} ({sum(nights_by_month.values()):.0f} nuits)")
        except Exception as e:
            errors.append(f"Pickup {pf.name} : {e}")

    if payload["pickupSnapshots"]:
        payload["pickupSnapshots"].sort(key=lambda s: s["exportDate"])
        print(f"  → {len(payload['pickupSnapshots'])} snapshots pickup chargés")

    # ── Résumé erreurs ──
    if errors:
        print("\n⚠ Avertissements :")
        for e in errors:
            print(f"  - {e}")

    # ── Génération HTML ──
    export_date_str = datetime.now().strftime("%d/%m/%Y à %H:%M")
    inject_into_html(template, output, payload, export_date_str)


if __name__ == "__main__":
    main()
